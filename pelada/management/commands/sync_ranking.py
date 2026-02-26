from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from pelada.models import PlayerStat


def to_int(value, field_name: str) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        raise CommandError(f"Valor inválido para {field_name}: {value!r}")


def to_decimal_pct(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")

    if isinstance(value, str):
        value = value.strip().replace("%", "").replace(",", ".")

    try:
        d = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise CommandError(f"Valor inválido para pct_vitorias: {value!r}")

    # Se vier 0.62 achando que é 62%
    if d <= 1 and d != 0:
        d = d * Decimal("100")

    if d < 0:
        d = Decimal("0")
    if d > 100:
        d = Decimal("100")

    return d.quantize(Decimal("0.01"))


class Command(BaseCommand):
    help = "Sincroniza o ranking a partir de um arquivo .xlsx (aba 'ranking')."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Caminho do arquivo .xlsx (ex: C:\\\\pelada\\\\pelada_ranking.xlsx)",
        )
        parser.add_argument("--sheet", default="ranking", help="Nome da aba (padrão: ranking)")
        parser.add_argument(
            "--clear-missing",
            action="store_true",
            help="Remove do banco quem NÃO estiver na planilha (use com cuidado).",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser().resolve()
        sheet_name = options["sheet"]
        clear_missing = options["clear_missing"]

        if not file_path.exists():
            raise CommandError(f"Arquivo não encontrado: {file_path}")

        wb = load_workbook(filename=str(file_path), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {wb.sheetnames}")

        ws = wb[sheet_name]

        # Cabeçalho (linha 1)
        header = []
        for cell in ws[1]:
            header.append((str(cell.value).strip().lower() if cell.value else ""))

        required = ["nome", "gols", "partidas_jogadas", "partidas_vencidas", "pct_vitorias"]
        for col in required:
            if col not in header:
                raise CommandError(f"Coluna obrigatória ausente: '{col}'. Cabeçalho encontrado: {header}")

        idx = {name: header.index(name) for name in required}

        created = 0
        updated = 0
        skipped = 0
        nomes_na_planilha = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            nome_raw = row[idx["nome"]]
            if nome_raw is None or str(nome_raw).strip() == "":
                skipped += 1
                continue

            nome = str(nome_raw).strip()
            nomes_na_planilha.add(nome.lower())

            gols = to_int(row[idx["gols"]], "gols")
            partidas_jogadas = to_int(row[idx["partidas_jogadas"]], "partidas_jogadas")
            partidas_vencidas = to_int(row[idx["partidas_vencidas"]], "partidas_vencidas")
            pct_vitorias = to_decimal_pct(row[idx["pct_vitorias"]])

            if partidas_vencidas > partidas_jogadas:
                raise CommandError(
                    f"Inconsistência para '{nome}': partidas_vencidas ({partidas_vencidas}) > partidas_jogadas ({partidas_jogadas})"
                )

            _, is_created = PlayerStat.objects.update_or_create(
                nome=nome,
                defaults={
                    "gols": gols,
                    "partidas_jogadas": partidas_jogadas,
                    "partidas_vencidas": partidas_vencidas,
                    "pct_vitorias": pct_vitorias,
                },
            )

            if is_created:
                created += 1
            else:
                updated += 1

        removed = 0
        if clear_missing:
            for obj in PlayerStat.objects.all():
                if obj.nome.lower() not in nomes_na_planilha:
                    obj.delete()
                    removed += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Sync concluído. created={created} updated={updated} skipped={skipped} removed={removed} file={file_path.name}"
            )
        )
