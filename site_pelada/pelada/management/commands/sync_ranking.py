from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from pathlib import Path
import openpyxl

from pelada.models import PlayerStat


class Command(BaseCommand):
    help = "Sincroniza o ranking a partir da planilha Excel (modo espelho: apaga tudo e recria)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Caminho para o arquivo .xlsx",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])

        if not file_path.exists():
            raise CommandError(f"Arquivo não encontrado: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Erro ao abrir Excel: {e}")

        ws = wb.worksheets[0]

        header = [
            str(c.value).strip().lower() if c.value else ""
            for c in ws[1]
        ]

        def find_col(name):
            return header.index(name) if name in header else None

        col_nome = find_col("nome")
        col_gols = find_col("gols")
        col_jogos = find_col("partidas_jogadas")
        col_vit = find_col("partidas_vencidas")
        col_pct = find_col("pct_vitorias")

        if None in [col_nome, col_gols, col_jogos, col_vit, col_pct]:
            raise CommandError("Verifique os nomes das colunas no Excel.")

        novos_registros = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_nome]:
                continue

            novos_registros.append(
                PlayerStat(
                    nome=str(row[col_nome]).strip(),
                    gols=int(row[col_gols] or 0),
                    partidas_jogadas=int(row[col_jogos] or 0),
                    partidas_vencidas=int(row[col_vit] or 0),
                    pct_vitorias=float(row[col_pct] or 0),
                )
            )

        # 🔥 ESPELHO TOTAL
        with transaction.atomic():
            PlayerStat.objects.all().delete()
            PlayerStat.objects.bulk_create(novos_registros)

        self.stdout.write(
            self.style.SUCCESS(
                f"Sincronização concluída. {len(novos_registros)} jogadores importados."
            )
        )